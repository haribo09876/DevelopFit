from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from openpyxl import Workbook
import re
import time
import sys

# # movieCdList 저장
# read_xlsx = load_workbook(
#     r'C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieCdTitle.xlsx')
# read_sheet = read_xlsx.active

# movieCd_col = read_sheet['A']
# movieCdList = []
# for cell in movieCd_col:
#     movieCdList.append(cell.value)


# # movieNmList 저장
# read_xlsx = load_workbook(
#     r'C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieCdTitle.xlsx')
# read_sheet = read_xlsx.active

# movieNm_col = read_sheet['B']
# movieNmList = []
# for cell in movieNm_col:
#     movieNmList.append(cell.value)


# # movieSummaryInfo 저장
# read_xlsx = load_workbook(
#     r'C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieDbSecond.xlsx')
# read_sheet = read_xlsx.active

# movieSummaryInfo_col = read_sheet['E']
# sentences = []
# for cell in movieSummaryInfo_col:
#     sentences.append(cell.value)


# # <영화 정보 저장>
# wbFirst = Workbook()
# wsFirst = wbFirst.active
# wsFirst.append(['영화코드', '영화제목', '메인포스터', '시놉시스', '요약정보', '상영시간'])

# browser = webdriver.Chrome()
# for n in range(1285, 3074):
#     # 브라우저 자동 꺼짐 방지시 사용
#     # chrome_options = Options()
#     # chrome_options.add_experimental_option("detach", True)
#     # chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
#     # browser = webdriver.Chrome(options=chrome_options)

#     # 영화 검색 페이지로 이동
#     browser.get(
#         "https://www.kobis.or.kr/kobis/business/mast/mvie/searchUserMovCdList.do")

#     # 영화 코드로 검색
#     try:
#         inoutObj = browser.find_element(By.NAME, "movieCd")
#         inoutObj.send_keys(movieCdList[n])
#         searchButtonObj = browser.find_element(
#             By.XPATH, "//*[@id='searchForm']/div[1]/div[7]/button")
#         searchButtonObj.click()
#     except:
#         pass

#     # 영화 제목
#     try:
#         titleObj = browser.find_element(
#             By.XPATH, "//*[@id='content']/div[4]/table/tbody/tr[1]/td[1]/a")
#         title = titleObj.text
#     except:
#         pass

#     # 영화 상세 페이지로 이동
#     try:
#         toDatailObj = browser.find_element(
#             By.XPATH, "//*[@id='content']/div[4]/table/tbody/tr[1]/td[1]/a")
#         toDatailObj.click()
#     except:
#         pass

#     time.sleep(1)
#     # 메인포스터
#     try:
#         mainPosterObj = browser.find_element(
#             By.XPATH, "//*[@id='ui-id-1']/div/div[1]/div[2]/a")
#         mainPoster = mainPosterObj.get_attribute("href")
#     except:
#         pass

#     # 시놉시스
#     try:
#         synopsisObj = browser.find_element(By.CLASS_NAME, "desc_info")
#         synopsis = synopsisObj.text
#     except:
#         pass

#     # 요약정보
#     try:
#         summaryInfoObj = browser.find_element(
#             By.XPATH, "//*[@id='ui-id-1']/div/div[1]/div[2]/dl/dd[4]")
#         summaryInfo = summaryInfoObj.text
#     except:
#         pass

#     # 상영시간
#     try:
#         runningTimeArr = []
#         runningTimeArr = summaryInfo.split('분')
#         runningTimeFront = runningTimeArr[0]
#         runnigTime = re.sub(r'[^0-9]', '', runningTimeFront)
#     except:
#         pass
#     time.sleep(1)

#     print(movieCdList[n], "", title, "", mainPoster,
#           "", synopsis, "", summaryInfo, "", runnigTime)
#     time.sleep(1)
#     wsFirst.append([movieCdList[n], title, mainPoster,
#                     synopsis, summaryInfo, runnigTime])
#     wbFirst.save('C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieInfo5.xlsx')
#     time.sleep(1)
# wbFirst.close()


# # <영화 예고편 링크 저장>
# wbFirst = Workbook()
# wsFirst = wbFirst.active
# wsFirst.append(['영화제목', '예고편링크'])

# browser = webdriver.Chrome()

# for n in range(1, 3074):
#     browser.get("https://www.youtube.com/results?search_query=영화+" + movieNmList[n] + "+예고편")

#     # 영화 예고편 상세
#     try:
#         movieTrailerObj = browser.find_element(By.XPATH, "/html/body/ytd-app/div[1]/ytd-page-manager/ytd-search/div[1]/ytd-two-column-search-results-renderer/div/ytd-section-list-renderer/div[2]/ytd-item-section-renderer[2]/div[3]/ytd-video-renderer[1]/div[1]/ytd-thumbnail/a")
#         movieTrailerLink = movieTrailerObj.get_attribute("href")
#     except:
#         pass
#     time.sleep(1)
#     print(movieNmList[n], "", movieTrailerLink)
#     wsFirst.append([movieNmList[n], movieTrailerLink])
#     wbFirst.save('C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieTrailer1.xlsx')
#     time.sleep(1)
# wbFirst.close()


# # <영화 관람가 파싱 및 저장>
# wbFirst = Workbook()
# wsFirst = wbFirst.active
# wsFirst.append(['관람가'])

# # 정규표현식을 사용하여 '관람'이 포함된 단어 추출
# pattern = r'\b\w*관람\w*\b'

# for sentence in sentences:
#     match = re.search(pattern, sentence)
#     if match:
#         print(match.group())
#         wsFirst.append([match.group()])
#     else:
#         print("")
#         wsFirst.append([""])
# wbFirst.save('C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieInfo5.xlsx')
# wbFirst.close()


# # <영화 사이드포스터, 스틸컷, 주연배우 저장>
# wbFirst = Workbook()
# wsFirst = wbFirst.active
# wsFirst.append(['영화코드', '사이드포스터'])

# wbSecond = Workbook()
# wsSecond = wbSecond.active
# wsSecond.append(['영화코드', '스틸컷'])

# wbThird = Workbook()
# wsThird = wbThird.active
# wsThird.append(['영화코드', '주연배우'])

# browser = webdriver.Chrome()

# for n in range(2180, 3074):
#     browser.get(
#         "https://www.kobis.or.kr/kobis/business/mast/mvie/searchUserMovCdList.do")

#     # 영화 코드로 검색 및 상세 페이지로 이동
#     inoutObj = browser.find_element(By.NAME, "movieCd")
#     inoutObj.send_keys(movieCdList[n])
#     searchButtonObj = browser.find_element(
#         By.XPATH, "//*[@id='searchForm']/div[1]/div[7]/button")
#     searchButtonObj.click()

#     toDatailObj = browser.find_element(
#         By.XPATH, "//*[@id='content']/div[4]/table/tbody/tr[1]/td[1]/a")
#     toDatailObj.click()
#     time.sleep(1)

#     # <영화 사이드포스터>
#     try:
#         for m in range(1, 11):
#             sidePosterObjs = browser.find_element(
#                 By.XPATH, "/html/body/div[2]/div[2]/div/div[1]/div[3]/div[1]/div/div/div/div[" + str(m) + "]/a/img")
#             sidePoster = sidePosterObjs.get_attribute("src")
#             if sidePoster:
#                 print(movieCdList[n], "", sidePoster)
#                 wsFirst.append([movieCdList[n], sidePoster])
#         for l in range(1, 21):
#             stillImageObjs = browser.find_element(
#                 By.XPATH, "/html/body/div[2]/div[2]/div/div[1]/div[4]/div[1]/div/div[1]/div/div[" + str(l) + "]/a/img")
#             stillImage = stillImageObjs.get_attribute("src")
#             if stillImage:
#                 print(movieCdList[n], "", stillImage)
#                 wsSecond.append([movieCdList[n], stillImage])
#         for k in range(1, 11):
#             castObj = browser.find_element(
#                 By.XPATH, "/html/body/div[2]/div[2]/div/div[1]/div[6]/div/dl/div[2]/dd/table/tbody/tr/td/a[" + str(k) + "]")
#             cast = castObj.text
#             if cast:
#                 print(movieCdList[n], "", cast)
#                 wsThird.append([movieCdList[n], cast])
#     except:
#         pass
#     wbFirst.save(
#         'C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieSidePoster3.xlsx')
#     wbFirst.close()
#     wbSecond.save(
#         'C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieStillImage3.xlsx')
#     wbSecond.close()
#     wbThird.save('C:/Users/ErYK/Desktop/Coding/팀 프로젝트/DB추가/movieCast3.xlsx')
#     wbThird.close()
#     time.sleep(1)
